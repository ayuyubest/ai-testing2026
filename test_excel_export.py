"""测试Excel导出功能"""
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 测试数据
testcases = [
    {
        "用例编号": "TC-TEST-001",
        "用例标题": "用户登录-正常登录",
        "前置条件": "用户已注册且账号状态正常",
        "测试步骤": "1. 打开登录页面\n2. 输入正确的用户名和密码\n3. 点击登录按钮",
        "预期结果": "登录成功，跳转到首页",
        "优先级": "P0",
        "用例类型": "功能测试"
    },
    {
        "用例编号": "TC-TEST-002",
        "用例标题": "用户登录-密码错误",
        "前置条件": "用户已注册",
        "测试步骤": "1. 打开登录页面\n2. 输入正确的用户名和错误的密码\n3. 点击登录按钮",
        "预期结果": "提示密码错误，不允许登录",
        "优先级": "P1",
        "用例类型": "功能测试"
    }
]

# 创建输出目录
output_dir = Path("test_output")
output_dir.mkdir(exist_ok=True)

# 生成文件名
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filepath = output_dir / f"测试用例_{timestamp}.xlsx"

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "测试用例"

# 表头
headers = ["用例编号", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "用例类型"]
border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# 写入数据
for row_idx, tc in enumerate(testcases, 2):
    ws.cell(row=row_idx, column=1, value=tc.get("用例编号", "")).border = border
    ws.cell(row=row_idx, column=2, value=tc.get("用例标题", "")).border = border
    ws.cell(row=row_idx, column=3, value=tc.get("前置条件", "")).border = border
    ws.cell(row=row_idx, column=4, value=tc.get("测试步骤", "")).border = border
    ws.cell(row=row_idx, column=5, value=tc.get("预期结果", "")).border = border
    ws.cell(row=row_idx, column=6, value=tc.get("优先级", "")).border = border
    ws.cell(row=row_idx, column=7, value=tc.get("用例类型", "")).border = border

# 调整列宽
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12

# 保存
wb.save(filepath)
print(f"✅ 成功保存 {len(testcases)} 条测试用例到 {filepath}")

