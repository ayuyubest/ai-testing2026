"""
测试用例Excel导出工具
"""
from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from langchain_core.tools import tool


@tool
def save_testcases_to_excel(
    testcases: List[Dict[str, Any]],
    filename: str = None,
    sheet_name: str = "测试用例"
) -> str:
    """
    将测试用例保存到Excel文件，默认保存到test_output文件夹

    Args:
        testcases: 测试用例列表，每个用例包含字段如: 用例编号、用例标题、前置条件、测试步骤、预期结果、优先级等
        filename: Excel文件名（不含路径），如果不提供则自动生成带时间戳的文件名
        sheet_name: 工作表名称，默认为"测试用例"

    Returns:
        str: 保存结果消息
    """
    if not testcases:
        return "错误：测试用例列表为空"

    # 确定项目根目录（从当前文件向上3级）
    project_root = Path(__file__).parent.parent.parent.parent
    output_dir = project_root / "test_output"
    output_dir.mkdir(parents=True, exist_ok=True)

    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"测试用例_{timestamp}.xlsx"
    elif not filename.endswith('.xlsx'):
        filename = f"{filename}.xlsx"

    filepath = output_dir / filename
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # 定义表头
    headers = ["用例编号", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "用例类型"]
    
    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row_idx, testcase in enumerate(testcases, 2):
        ws.cell(row=row_idx, column=1, value=testcase.get("用例编号", "")).border = border
        ws.cell(row=row_idx, column=2, value=testcase.get("用例标题", "")).border = border
        ws.cell(row=row_idx, column=3, value=testcase.get("前置条件", "")).border = border
        ws.cell(row=row_idx, column=4, value=testcase.get("测试步骤", "")).border = border
        ws.cell(row=row_idx, column=5, value=testcase.get("预期结果", "")).border = border
        ws.cell(row=row_idx, column=6, value=testcase.get("优先级", "")).border = border
        ws.cell(row=row_idx, column=7, value=testcase.get("用例类型", "")).border = border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    
    # 保存文件
    wb.save(filepath)

    return f"✅ 成功保存 {len(testcases)} 条测试用例到 {filepath}"

